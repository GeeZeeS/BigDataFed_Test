from openpyxl import load_workbook
import json
import re
import os
from os.path import exists

from datetime import datetime


final = []


def date_formatter(date_string):
    date_format = '%m-%d-%Y'
    formated = datetime.strptime(date_string, date_format).strftime('%Y-%m-%d')
    return formated


def load_json(data):
    json_url = 'data.json'
    if exists(json_url):
        data = json.load(open(json_url))
    else:
        generator()
        data = json.load(open(json_url))
    return data


###############################################################################
# Parse Dates From String
###############################################################################


def get_dates(from_to_str, obj):
    processed_data = re.findall(r'(\w+) *:(?: *([\w.-]+))?', from_to_str)
    from_date = processed_data[0][1]
    to_date = processed_data[1][1]
    obj['week_start'] = date_formatter(from_date)
    obj['week_end'] = date_formatter(to_date)
    return obj


###############################################################################
# Get values from excel page for dates and and data
###############################################################################


def get_data(working_sheet, obj):
    for i in range(7, working_sheet.max_row + 1):
        cell = f'A{i}'
        name = working_sheet[cell].value
        if name is None:
            break
        else:
            value = working_sheet[f'C{i}'].value
            obj[f'{name}'] = value
    return obj


###############################################################################
# Process Excel file, Get pages, Run functions for parsing data and write data to JSON
###############################################################################


def process_workbook(workbook):
    sheets = workbook.sheetnames
    for sheet_name in sheets:
        fin = {}
        working_sheet = workbook[f'{sheet_name}']
        working_data = working_sheet['A3'].value
        fin = get_dates(working_data, fin)
        fin = get_data(working_sheet, fin)
        final.append(fin)


###############################################################################
# Generate JSON File from excel files inside xlsx folder
###############################################################################


def generator():
    path = 'xlsx/'
    files = os.listdir(path)
    files_xls = [f for f in files if f[-4:] == 'xlsx']
    for file in files_xls:
        wb = load_workbook(f'xlsx/{file}')
        process_workbook(wb)
    save_file()


###############################################################################
# Save Data to JSON File
###############################################################################


def save_file():
    with open('data.json', 'w') as f:
        json.dump(final, f, indent=4)
